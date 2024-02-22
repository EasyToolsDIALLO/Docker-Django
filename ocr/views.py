from django.shortcuts import render, get_object_or_404, redirect
import io
import openpyxl
from django.http import HttpResponse
from django.template import loader
from .models import *
from .forms import MyImageForm
import pytesseract
import cv2
import numpy
import json
import pprintpp

# Create your views here.


def home(request):
    return render(request, 'home.html')


def export_to_excel(request, pk):
    reload = 0
    if reload == 0:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="facture.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Facture'

        # Write header row
        header = ['Nom', 'Quantité', 'Prix Unitaire', 'Montant', 'Id Facture']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        id_facture = get_object_or_404(factures, pk=pk)
        facture = factures.objects.get(pk=id_facture)
        mes_produits = produit.objects.filter(codeProduit=id_facture).values_list(
            'nom', 'quantite', 'prix_unitaire', 'somme', 'codeProduit')

        for row_num, row in enumerate(mes_produits, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        reload = 1
        return response
    else:
        return redirect("")


def les_factures(request):
    if request.method == 'POST':
        form = MyImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        image_url = "media/" + str(MyImageModel. objects. last().image)
        une_facture = get_facture(image_url)
        facture_, produits = enregistrer(une_facture)

        les_factures = factures.objects.get(pk=facture_.facture_ID)
        les_produits = produit.objects.filter(codeProduit=facture_.facture_ID)
        image = MyImageModel.objects.all().last()
        print(image.image)

        facture = factures.objects.all()
        somme = 0
        for f in facture:
            somme += float(f.total)

        return render(request, 'les_factures.html',
                      {
                          'factures': facture,
                          'factures_total': somme

                      })

    else:
        form = MyImageForm()
        facture = factures.objects.all()
        somme = 0
        for f in facture:
            somme += float(f.total)
        return render(request, 'les_factures.html',  {'form': form, 'factures': facture, 'factures_total': somme})


def traitement(request, pk):
    facture_ID = get_object_or_404(factures, pk=pk)
    les_factures = factures.objects.get(pk=facture_ID)
    les_produits = produit.objects.filter(codeProduit=facture_ID)

    return render(request, 'get_facture.html', {'les_factures': les_factures, 'les_produits': les_produits})


def enregistrer(la_facture):
    facture_ID = la_facture['id']
    type_facture = la_facture['type']
    emetteur = la_facture['emetteur']
    client = la_facture['recepteur']
    mail_emetteur = la_facture['mail']
    date_emission = la_facture['date_emission']
    date_echeance = la_facture['date_echeance']
    conditions = la_facture['conditions']
    total = la_facture['total']

    facture_ = factures(
        facture_ID=facture_ID,
        type_facture=type_facture,
        emetteur=emetteur,
        client=client,
        mail_emetteur=mail_emetteur,
        date_echeance=date_echeance,
        date_emission=date_emission,
        conditions=conditions,
        total=total,
        facture_image=MyImageModel.objects.all().last()
    )
    facture_.save()

    for i in range(0, len(la_facture['produits'])):

        nom = la_facture['produits'][i]['nom']
        quantite = la_facture['produits'][i]['quantite']
        prix_unitaire = la_facture['produits'][i]['prix_unitaire']
        somme = la_facture['produits'][i]['sous-total']
        produits = produit(codeProduit=facture_,
                           nom=nom, quantite=quantite,
                           prix_unitaire=prix_unitaire,
                           somme=somme)
        produits.save()
    # pprintpp.pprint(facture)
    return facture_, produits


def get_facture(url):
    img = cv2.imread(url)
    text = pytesseract.image_to_string(img)

    table = text.split('\n')
    table = list(filter(lambda x: x != "", table))
    numpy_table = numpy.array(table)
    # pprintpp.pprint(table)

    produit = numpy_table[12:-8]
    sous_total = numpy_table[-8].split()[1].replace(",", "")
    total = numpy_table[-7].split()[1].replace(",", "").replace("XOF", "")

    produits = produit.tolist()
    pprintpp.pprint(produits)

    les_produits = []

    dic = {
        "nom": "",
        "quantite": 0,
        "prix_unitaire": 0,
        "sous-total": 0
    }

    # pprintpp.pprint(produits)
    for x in produits:
        x = x.split()
        quantite, prix_unitaire, soustotal = x[-3::]
        index = x[0]
        nom = " ".join([str(item) for item in x[1:-3]])
        dic["nom"] = nom
        dic["quantite"] = float(quantite)
        dic["prix_unitaire"] = float(prix_unitaire.replace(',', ''))
        dic["sous-total"] = float(soustotal.replace(',', ''))

        les_produits.append(dic)

        dic = {}

    # print("-"*30)
    # print("\n\n")
    id = numpy_table[1]
    type_facture = numpy_table[0]
    debiteur = numpy_table[4].split()[-1]
    recepteur = numpy_table[5]
    mail_debiteur = numpy_table[7]
    date_emission = numpy_table[8].replace("Date de facture : ", "")
    conditions = numpy_table[9].replace("Conditions : ", "")
    date_echeance = numpy_table[10].replace("Date d’échéance : ", "")
    pays = numpy_table[6]

    facture = {
        "id": id,
        "type": type_facture,
        "emetteur": debiteur,
        "recepteur": recepteur,
        "mail": mail_debiteur,
        "date_emission": date_emission,
        "date_echeance": date_echeance,
        "conditions": conditions,
        "produits": les_produits,
        "sous_total": sous_total,
        "total": total,
        "pays": pays
    }

    return facture
